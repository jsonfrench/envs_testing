import requests
from bs4 import BeautifulSoup
import openpyxl

workbook = openpyxl.load_workbook("D:\\VSCode\\SRP\\nih_data.xlsx")
sheet_name = "SRP Project Data"
sheet = workbook[sheet_name]

foa_column = "C"
title_text = "FOA URL (use for beautifulsoup scraper)"

counter = 2

sheet.cell(row=1, column=53).value = "Full Text of Announcement"

for cell in sheet[foa_column]:
    if cell.value != title_text:
        url = cell.value
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        target_elements = soup.find_all('div', attrs={'data-section-code': 'FOD'})
        all_words = []
        for element in target_elements:
            words = [text.strip() for text in element.stripped_strings]
            all_words.extend(words)
        full_text = '\n'.join(all_words)

        print(full_text)

        sheet.cell(row=counter, column=53).value = full_text
        counter += 1
        
        workbook.save("D:\\VSCode\\SRP\\nih_data.xlsx")



import requests
from bs4 import BeautifulSoup
import openpyxl

workbook = openpyxl.load_workbook("D:\\VSCode\\SRP\\NIHProjectData.xlsx")
sheet_name = "ProjectData"
sheet = workbook[sheet_name]

foa_column = "C"
title_text = "FOA URL (use for beautifulsoup scraper)"
text_column = 55

counter = 2

sheet.cell(row=1, column=text_column).value = "Full Text of Announcement"

for cell in sheet[foa_column]:
    if cell.value != title_text:
        url = cell.value

        #===
        page = requests.get(url)

        soup = BeautifulSoup(page.content, "html")
        q=soup.get_text() #all text for FOA

        #remove first occurances of these
        t = q.replace("Part 2. Full Text of Announcement","",1)
        t = t.replace("Section II. Award Information","",1)

        #reduce text to title, purpose, and full announcement text, 
        u=t[t.find("Funding Opportunity Title"):t.find("Activity Code")]
        v=t[t.find("Notice of Funding Opportunity Purpose"):t.find("Key Dates")]
        w=t[t.find("Part 2. Full Text of Announcement"):t.find("Section II. Award Information")]
        x = u+v+w 
        y = x.replace("\t", " ")
        z = y.replace("\n", " ") #you may need to remove additional \n 
        #===

        sheet.cell(row=counter, column=text_column).value = z
        counter += 1        

        workbook.save("D:\\VSCode\\SRP\\NIHProjectData.xlsx")
        
        print("iteration:", counter)

print("done!")


'''
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        target_elements = soup.find_all('div', attrs={'data-section-code': 'FOD'})
        all_words = []
        for element in target_elements:
            words = [text.strip() for text in element.stripped_strings]
            all_words.extend(words)
        full_text = '\n'.join(all_words)

        print(full_text)

        sheet.cell(row=counter, column=text_column).value = full_text
        counter += 1
        
        workbook.save("D:\\VSCode\\SRP\\nih_data.xlsx")



# Insert code to read all FOA links into a list and process the list
# FOAlinks.txt

#Start for loop to process url list (I tried a few)

#URL = "https://grants.nih.gov/grants/guide/rfa-files/RFA-AA-23-001.html"
URL = "https://grants.nih.gov/grants/guide/rfa-files/RFA-AA-21-007.html"

page = requests.get(URL)

soup = BeautifulSoup(page.content, "html")
q=soup.get_text() #all text for FOA

#remove first occurances of these
t = q.replace("Part 2. Full Text of Announcement","",1)
t = t.replace("Section II. Award Information","",1)

#reduce text to title, purpose, and full announcement text, 
u=t[t.find("Funding Opportunity Title"):t.find("Activity Code")]
v=t[t.find("Notice of Funding Opportunity Purpose"):t.find("Key Dates")]
w=t[t.find("Part 2. Full Text of Announcement"):t.find("Section II. Award Information")]
x = u+v+w 
y = x.replace("\t", " ")
z = y.replace("\n", " ") #you may need to remove additional \n 

#write this text z to a file that uses the FOA number as its name: example RFA-AA-23-001.txt
print(z) 

#End for loop
'''
